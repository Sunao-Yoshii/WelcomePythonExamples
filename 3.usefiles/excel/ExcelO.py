import openpyxl as px

# Excel を新規に作成
book = px.Workbook()
sheet = book.active

sheet['A1'] = 'Welcome to my broken show!'
book.save("sample.xlsx")

# Excel を開いて値の読み取り
book = px.load_workbook('sample.xlsx')

# Sheet 名参照
print(book.sheetnames)

# Sheet 選択
sheet = book['Sheet']

# セル読み
print(sheet['A1'].value)

